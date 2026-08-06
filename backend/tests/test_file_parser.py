from __future__ import annotations

import io

from app.services import file_parser


def test_parse_csv_joins_columns_with_pipe():
    csv_bytes = b"name,age\nAlice,30\nBob,25\n"

    result = file_parser.parse_csv(csv_bytes)

    assert "name | age" in result
    assert "Alice | 30" in result
    assert "Bob | 25" in result


def test_parse_csv_truncates_after_max_rows(monkeypatch):
    monkeypatch.setattr(file_parser, "_MAX_ROWS", 2)
    csv_bytes = b"\n".join(f"row{i}".encode() for i in range(5))

    result = file_parser.parse_csv(csv_bytes)

    assert "truncated after 2 rows" in result


def test_parse_csv_truncates_output_length(monkeypatch):
    monkeypatch.setattr(file_parser, "_MAX_CHARS", 10)
    csv_bytes = b"a,b,c,d,e,f,g,h,i,j,k,l,m,n"

    result = file_parser.parse_csv(csv_bytes)

    assert len(result) == 10


def test_parse_xlsx_extracts_sheet_rows():
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Header1", "Header2"])
    ws.append(["v1", "v2"])
    buf = io.BytesIO()
    wb.save(buf)

    result = file_parser.parse_xlsx(buf.getvalue())

    assert "[Sheet: Sheet1]" in result
    assert "Header1 | Header2" in result
    assert "v1 | v2" in result


def test_parse_xlsx_skips_fully_empty_rows():
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["a"])
    ws.append([None, None])
    ws.append(["b"])
    buf = io.BytesIO()
    wb.save(buf)

    result = file_parser.parse_xlsx(buf.getvalue())

    # The fully-empty middle row is skipped; only "a" and "b" rows remain.
    assert result == "[Sheet: Sheet]\na | \nb | "


def test_parse_pdf_extracts_text_per_page():
    import pypdf

    writer = pypdf.PdfWriter()
    writer.add_blank_page(width=72, height=72)
    buf = io.BytesIO()
    writer.write(buf)

    # A blank page has no extractable text, so parse_pdf should return an
    # empty string rather than raising.
    result = file_parser.parse_pdf(buf.getvalue())
    assert result == ""


def test_parse_document_dispatches_by_extension():
    assert file_parser.parse_document(b"a,b\n1,2\n", "data.csv") != ""
    assert file_parser.parse_document(b"a,b\n1,2\n", "DATA.CSV") != ""
    assert file_parser.parse_document(b"anything", "notes.txt") == ""


def test_parse_document_dispatches_xlsx_and_xls(monkeypatch):
    calls = []
    monkeypatch.setattr(file_parser, "parse_xlsx", lambda b: calls.append(b) or "parsed")

    assert file_parser.parse_document(b"x", "sheet.xlsx") == "parsed"
    assert file_parser.parse_document(b"x", "sheet.xls") == "parsed"
    assert len(calls) == 2


def test_parse_document_dispatches_pdf(monkeypatch):
    monkeypatch.setattr(file_parser, "parse_pdf", lambda b: "pdf-text")

    assert file_parser.parse_document(b"x", "report.pdf") == "pdf-text"


def test_parse_pdf_extracts_text_when_page_has_content():
    from unittest.mock import MagicMock, patch

    fake_page = MagicMock()
    fake_page.extract_text.return_value = "Hello from a real PDF page"
    fake_reader = MagicMock()
    fake_reader.pages = [fake_page]

    with patch("pypdf.PdfReader", return_value=fake_reader):
        result = file_parser.parse_pdf(b"fake-pdf-bytes")

    assert "Hello from a real PDF page" in result
    assert "[Page 1]" in result


def test_parse_xlsx_truncates_after_max_rows(monkeypatch):
    import openpyxl

    monkeypatch.setattr(file_parser, "_MAX_ROWS", 3)
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(10):
        ws.append([f"row{i}"])
    buf = __import__("io").BytesIO()
    wb.save(buf)

    result = file_parser.parse_xlsx(buf.getvalue())
    assert "truncated after 3 rows" in result
