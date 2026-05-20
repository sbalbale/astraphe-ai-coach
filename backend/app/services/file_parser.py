"""
Document parsing utilities for the ASTRAPE AI coach.
Supports PDF, CSV, and Excel (XLSX/XLS) file types.
"""
from __future__ import annotations

import csv
import io

_MAX_CHARS = 8000
_MAX_ROWS = 500


def parse_pdf(file_bytes: bytes) -> str:
    from pypdf import PdfReader
    reader = PdfReader(io.BytesIO(file_bytes))
    parts: list[str] = []
    for i, page in enumerate(reader.pages, 1):
        text = (page.extract_text() or "").strip()
        if text:
            parts.append(f"[Page {i}]\n{text}")
    return "\n\n".join(parts)[:_MAX_CHARS]


def parse_csv(file_bytes: bytes) -> str:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    lines: list[str] = []
    for i, row in enumerate(reader):
        if i >= _MAX_ROWS:
            lines.append(f"... (truncated after {_MAX_ROWS} rows)")
            break
        lines.append(" | ".join(str(c) for c in row))
    return "\n".join(lines)[:_MAX_CHARS]


def parse_xlsx(file_bytes: bytes) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= _MAX_ROWS:
                rows.append(f"... (truncated after {_MAX_ROWS} rows)")
                break
            if any(v is not None for v in row):
                rows.append(" | ".join("" if v is None else str(v) for v in row))
        if rows:
            parts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
    return "\n\n".join(parts)[:_MAX_CHARS]


def parse_document(file_bytes: bytes, filename: str) -> str:
    """Parse a supported document and return its text content."""
    name = (filename or "").lower()
    if name.endswith(".pdf"):
        return parse_pdf(file_bytes)
    elif name.endswith(".csv"):
        return parse_csv(file_bytes)
    elif name.endswith((".xlsx", ".xls")):
        return parse_xlsx(file_bytes)
    return ""
